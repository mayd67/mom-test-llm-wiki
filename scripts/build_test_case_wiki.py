from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = ROOT / 'raw' / '测试资料' / '01_测试用例'
WIKI_ROOT = ROOT / 'wiki'
PROCESS_DIR = WIKI_ROOT / '02_测试标准&模板' / '测试过程规范'
TEMPLATE_DIR = WIKI_ROOT / '02_测试标准&模板' / '可复用模板'
TODAY = date.today().isoformat()

FIELD_DESCRIPTIONS = {
    'ID': '历史导出用例唯一标识，仅在导出分册中出现。',
    '用例名称': '可直接被执行人员理解的用例标题，通常带 TC 前缀或场景描述。',
    '所属模块': '按 `/一级/二级/功能点` 路径组织的模块树，用于导入后落位。',
    '标签': '常用来挂接需求编号、需求主题或功能域，便于追溯。',
    '前置条件': '执行前必须满足的账号、数据、状态和权限准备。',
    '步骤描述': '可执行的操作步骤；多步场景可拆多行。',
    '预期结果': '每一步对应的结果校验点。',
    '编辑模式': '导出资料中以 `TEXT` 为主，模板样例展示 `STEP`，skill 文档默认建议 `TEST`。',
    '备注': '补充来源测试点、风险点、验证结果或待确认信息。',
    '责任人': '当前用例的编写/维护责任人。',
    '用例等级': '优先级分层，历史资料以 `P0`~`P3` 为主。',
    '用例状态': '用例当前状态，历史导出资料全部为 `未开始`。',
}


@dataclass
class ExistingPage:
    title: str
    path: Path
    text: str


@dataclass
class GeneratedPage:
    title: str
    path: Path
    page_type: str
    status: str
    tags: list[str]
    summary: str
    source: list[str]
    body: list[str]


def normalize(value: object) -> str:
    if value is None:
        return ''
    return str(value).strip()


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def parse_title(text: str, fallback: str) -> str:
    title_match = re.search(r'^title:\s*(.+?)\s*$', text, re.MULTILINE)
    if title_match:
        return title_match.group(1).strip().strip('"')
    heading_match = re.search(r'^#\s+(.+?)\s*$', text, re.MULTILINE)
    if heading_match:
        return heading_match.group(1).strip()
    return fallback


def load_existing_pages() -> list[ExistingPage]:
    pages: list[ExistingPage] = []
    for path in sorted(WIKI_ROOT.rglob('*.md')):
        if path.name in {'index.md', 'log.md'}:
            continue
        if '原始资料镜像' in path.parts:
            continue
        text = path.read_text(encoding='utf-8')
        pages.append(ExistingPage(parse_title(text, path.stem), path, text))
    return pages


def top_level_area(module_path: str) -> str:
    segments = [segment for segment in module_path.split('/') if segment]
    if len(segments) >= 2 and segments[0] == 'MOM产品功能用例':
        return segments[1]
    if segments:
        return segments[0]
    return '未分类'


def second_level_area(module_path: str) -> str:
    segments = [segment for segment in module_path.split('/') if segment]
    if len(segments) >= 3 and segments[0] == 'MOM产品功能用例':
        return segments[2]
    if len(segments) >= 2:
        return segments[1]
    if segments:
        return segments[0]
    return '未分类'


def counter_to_text(counter: Counter[str], limit: int = 8) -> str:
    if not counter:
        return '无'
    return '；'.join(f'{key} {value}' for key, value in counter.most_common(limit))


def table(lines: list[list[str]]) -> list[str]:
    if not lines:
        return ['- 空表']
    width = max(len(row) for row in lines)
    padded = [row + [''] * (width - len(row)) for row in lines]
    rendered = [
        '| ' + ' | '.join(cell.replace('|', '\\|') for cell in padded[0]) + ' |',
        '| ' + ' | '.join(['---'] * width) + ' |',
    ]
    for row in padded[1:]:
        rendered.append('| ' + ' | '.join(cell.replace('|', '\\|') for cell in row) + ' |')
    return rendered


def write_page(page: GeneratedPage) -> None:
    content = [
        '---',
        f'title: {page.title}',
        f'type: {page.page_type}',
        f'status: {page.status}',
        'tags:',
    ]
    for tag in page.tags:
        content.append(f'  - {tag}')
    content.extend(
        [
            f'summary: {page.summary}',
            'source:',
        ]
    )
    for item in page.source:
        content.append(f'  - {item}')
    content.extend([f'updated: {TODAY}', '---', ''])
    content.extend(page.body)
    page.path.parent.mkdir(parents=True, exist_ok=True)
    page.path.write_text('\n'.join(content).rstrip() + '\n', encoding='utf-8')


def replace_once(path: Path, old: str, new: str) -> None:
    text = path.read_text(encoding='utf-8')
    if new in text:
        return
    if old not in text:
        raise ValueError(f'Expected block not found in {rel(path)}')
    path.write_text(text.replace(old, new, 1), encoding='utf-8')


def append_section_if_missing(path: Path, section_title: str, lines: list[str]) -> None:
    text = path.read_text(encoding='utf-8')
    if section_title in text:
        return
    addition = '\n'.join(['', section_title, '', *lines]).rstrip() + '\n'
    path.write_text(text.rstrip() + addition, encoding='utf-8')


def find_related_titles(terms: list[str], existing_pages: list[ExistingPage], exclude: set[str] | None = None, limit: int = 8) -> list[str]:
    exclude = exclude or set()
    scored: dict[str, tuple[int, str]] = {}
    for term in terms:
        cleaned = term.strip()
        if not cleaned:
            continue
        variants = {cleaned}
        code_match = re.match(r'^(DNW\d+)', cleaned)
        if code_match:
            variants.add(code_match.group(1))
        if '-' in cleaned:
            _, suffix = cleaned.split('-', 1)
            if suffix:
                variants.add(suffix)
        for page in existing_pages:
            if page.title in exclude:
                continue
            score = 0
            for variant in variants:
                if variant in page.title:
                    score += 5
                elif variant in page.text:
                    score += 1
            if score == 0:
                continue
            if '需求拆解' in page.title:
                score += 4
            if '需求导航' in page.title or '需求地图' in page.title:
                score += 3
            if '产品资料' in page.title:
                score += 1
            current = scored.get(page.title)
            rel_path = rel(page.path)
            if current is None or score > current[0] or (score == current[0] and rel_path < current[1]):
                scored[page.title] = (score, rel_path)
    sorted_titles = sorted(scored.items(), key=lambda item: (-item[1][0], item[1][1], item[0]))
    return [title for title, _ in sorted_titles[:limit]]


def scan_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [normalize(value) for value in row]
        while values and values[-1] == '':
            values.pop()
        if values:
            rows.append(values)
    header = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    index = {name: position for position, name in enumerate(header)}

    def value_at(row: list[str], key: str) -> str:
        position = index.get(key)
        if position is None or position >= len(row):
            return ''
        return row[position]

    modules = Counter[str]()
    tags = Counter[str]()
    owners = Counter[str]()
    priorities = Counter[str]()
    statuses = Counter[str]()
    edit_modes = Counter[str]()
    top_areas = Counter[str]()
    second_areas = Counter[str]()
    module_paths: set[str] = set()
    tag_values: set[str] = set()
    sample_cases: list[tuple[str, str, str, str, str]] = []

    for row in data_rows:
        case_id = value_at(row, 'ID')
        case_name = value_at(row, '用例名称')
        module = value_at(row, '所属模块')
        tag = value_at(row, '标签')
        owner = value_at(row, '责任人')
        priority = value_at(row, '用例等级')
        status = value_at(row, '用例状态')
        edit_mode = value_at(row, '编辑模式')

        if module:
            modules[module] += 1
            module_paths.add(module)
            top_areas[top_level_area(module)] += 1
            second_areas[second_level_area(module)] += 1
        if tag:
            tags[tag] += 1
            tag_values.add(tag)
        if owner:
            owners[owner] += 1
        if priority:
            priorities[priority] += 1
        if status:
            statuses[status] += 1
        if edit_mode:
            edit_modes[edit_mode] += 1
        if case_name and len(sample_cases) < 5:
            sample_cases.append((case_id, case_name, module, tag, priority))

    return {
        'file': rel(path),
        'sheet': sheet.title,
        'header': header,
        'row_count': len(data_rows),
        'modules': modules,
        'tags': tags,
        'owners': owners,
        'priorities': priorities,
        'statuses': statuses,
        'edit_modes': edit_modes,
        'top_areas': top_areas,
        'second_areas': second_areas,
        'module_count': len(module_paths),
        'tag_count': len(tag_values),
        'sample_cases': sample_cases,
        'module_paths': module_paths,
        'tag_values': tag_values,
    }


def build_case_book_page(title: str, path: Path, workbook: dict[str, object], existing_pages: list[ExistingPage], overview_title: str) -> GeneratedPage:
    tags = list(workbook['tags'].keys())[:6]
    related_titles = find_related_titles(tags, existing_pages, exclude={title, overview_title})
    header_rows = [[header, FIELD_DESCRIPTIONS.get(header, '字段说明待补充。')] for header in workbook['header']]
    header_table = table([['列名', '说明'], *header_rows])
    sample_rows = []
    for case_id, case_name, module, tag, priority in workbook['sample_cases']:
        sample_rows.append([
            case_id or '—',
            case_name,
            module or '—',
            tag or '—',
            priority or '—',
        ])
    body = [
        f'# {title}',
        '',
        '## 资料概览',
        '',
        f'- 来源文件：`{workbook["file"]}`',
        f'- 工作表：`{workbook["sheet"]}`',
        f'- 用例规模：{workbook["row_count"]} 条历史用例，覆盖 {workbook["module_count"]} 条模块路径、{workbook["tag_count"]} 个需求标签。',
        f'- 负责人分布：{counter_to_text(workbook["owners"], 6)}。',
        f'- 优先级分布：{counter_to_text(workbook["priorities"], 6)}。',
        f'- 状态分布：{counter_to_text(workbook["statuses"], 6)}。',
        f'- 编辑模式：{counter_to_text(workbook["edit_modes"], 6)}。',
        '',
        '## 一级业务域分布',
        '',
    ]
    body.extend(table([['一级业务域', '用例数'], *[[name, str(count)] for name, count in workbook['top_areas'].most_common(12)]]))
    body.extend([
        '',
        '## 高频模块路径',
        '',
    ])
    body.extend(table([['模块路径', '用例数'], *[[name, str(count)] for name, count in workbook['modules'].most_common(12)]]))
    body.extend([
        '',
        '## 高频需求标签',
        '',
    ])
    body.extend(table([['标签', '用例数'], *[[name, str(count)] for name, count in workbook['tags'].most_common(12)]]))
    body.extend([
        '',
        '## 字段结构',
        '',
    ])
    body.extend(header_table)
    body.extend([
        '',
        '## 示例用例',
        '',
    ])
    body.extend(table([['ID', '用例名称', '所属模块', '标签', '优先级'], *sample_rows]))
    body.extend([
        '',
        '## 关联页面',
        '',
        f'- [[{overview_title}]]',
        '- [[测试用例资料总览]]',
        '- [[测试用例设计]]',
        '- [[测试用例规范]]',
        '- [[测试用例模板]]',
    ])
    for related_title in related_titles[:6]:
        body.append(f'- [[{related_title}]]')

    return GeneratedPage(
        title=title,
        path=path,
        page_type='asset',
        status='active',
        tags=['testing', 'mom', 'case-design'],
        summary=f'基于 {workbook["file"]} 整理的 MOM 历史测试用例分册摘要，沉淀覆盖域、标签和追溯入口。',
        source=[workbook['file']],
        body=body,
    )


def main() -> None:
    existing_pages = load_existing_pages()

    export_files = sorted((RAW_ROOT / '01_MOM产品测试用例').glob('*.xlsx'))
    template_file = RAW_ROOT / '03_测试用例导入模板' / '测试用例模版 .xlsx'
    sample_dir = RAW_ROOT / '02_测试用例样例'
    skill_root = RAW_ROOT / '生成测试用例skill'

    export_books = [scan_workbook(path) for path in export_files]
    template_book = scan_workbook(template_file)

    total_cases = sum(book['row_count'] for book in export_books)
    total_modules: set[str] = set()
    total_tags = Counter[str]()
    total_owners = Counter[str]()
    total_priorities = Counter[str]()
    total_statuses = Counter[str]()
    total_edit_modes = Counter[str]()
    total_areas = Counter[str]()
    for book in export_books:
        total_modules.update(book['module_paths'])
        total_tags.update(book['tags'])
        total_owners.update(book['owners'])
        total_priorities.update(book['priorities'])
        total_statuses.update(book['statuses'])
        total_edit_modes.update(book['edit_modes'])
        total_areas.update(book['top_areas'])

    overall_related = find_related_titles(list(total_tags.keys())[:12], existing_pages, exclude={'测试用例资料总览', 'MOM产品测试用例总览'})

    workbook_titles = [f'MOM产品测试用例分册{index}' for index in range(1, len(export_books) + 1)]
    workbook_pages = []
    for index, workbook in enumerate(export_books, start=1):
        workbook_pages.append(
            build_case_book_page(
                title=f'MOM产品测试用例分册{index}',
                path=PROCESS_DIR / f'MOM产品测试用例分册{index}.md',
                workbook=workbook,
                existing_pages=existing_pages,
                overview_title='MOM产品测试用例总览',
            )
        )

    overview_table = [['分册', '用例数', '一级业务域热点', '标签热点', '结构页']]
    for title, workbook in zip(workbook_titles, export_books, strict=True):
        overview_table.append([
            title,
            str(workbook['row_count']),
            '；'.join(f'{name} {count}' for name, count in workbook['top_areas'].most_common(3)),
            '；'.join(f'{name} {count}' for name, count in workbook['tags'].most_common(3)),
            f'[[{title}]]',
        ])

    root_page = GeneratedPage(
        title='测试用例资料总览',
        path=PROCESS_DIR / '测试用例资料总览.md',
        page_type='asset',
        status='active',
        tags=['testing', 'case-design', 'process'],
        summary='统一汇总 raw/测试资料/01_测试用例 下的 MOM 产品测试用例导出、导入模板、样例占位和生成 skill 资料。',
        source=['raw/测试资料/01_测试用例'],
        body=[
            '# 测试用例资料总览',
            '',
            '## 资料入口',
            '',
            '- [[MOM产品测试用例总览]]：5 个历史导出分册，合计 4399 条 MOM 产品测试用例。',
            '- [[测试用例导入模板说明]]：对齐本地 Excel 导入模板的列结构、录入方式和使用要点。',
            '- [[测试用例样例说明]]：记录 `02_测试用例样例` 目录当前为空的状态与后续补充建议。',
            '- [[测试用例生成Skill资料包]]：把需求分析、测试点拆分和模板化用例生成串成一条链路。',
            '',
            '## 规模概览',
            '',
            f'- 历史导出分册：{len(export_books)} 个工作簿，合计 {total_cases} 条用例，覆盖 {len(total_modules)} 条模块路径。',
            f'- 负责人分布：{counter_to_text(total_owners, 6)}。',
            f'- 一级业务域热点：{counter_to_text(total_areas, 8)}。',
            f'- 高频需求标签：{counter_to_text(total_tags, 10)}。',
            f'- 编辑模式现状：历史导出以 {counter_to_text(total_edit_modes, 4)} 为主，导入模板样例展示 `STEP`，skill 文档默认建议 `TEST`，导入前需统一平台约束。',
            '',
            '## 关键链接',
            '',
            '- [[测试标准与模板总览]]',
            '- [[测试过程规范总览]]',
            '- [[测试资产库总览]]',
            '- [[测试用例设计]]',
            '- [[测试用例规范]]',
            '- [[测试用例模板]]',
            '- [[核心业务需求导航]]',
            '- [[后台管理需求导航]]',
            '- [[智能能力需求导航]]',
        ] + [f'- [[{title}]]' for title in overall_related[:6]],
    )

    mom_overview_page = GeneratedPage(
        title='MOM产品测试用例总览',
        path=PROCESS_DIR / 'MOM产品测试用例总览.md',
        page_type='asset',
        status='active',
        tags=['testing', 'mom', 'case-design'],
        summary='按分册整理 MOM 产品历史测试用例导出资料，建立用例规模、模块覆盖、需求追踪和标准页之间的导航。',
        source=[book['file'] for book in export_books],
        body=[
            '# MOM产品测试用例总览',
            '',
            '## 总体概况',
            '',
            f'- 来源目录：`{rel(RAW_ROOT / "01_MOM产品测试用例")}`。',
            f'- 分册数量：{len(export_books)}。',
            f'- 用例总数：{total_cases}。',
            f'- 模块路径覆盖：{len(total_modules)} 条。',
            f'- 负责人分布：{counter_to_text(total_owners, 6)}。',
            f'- 优先级分布：{counter_to_text(total_priorities, 6)}。',
            f'- 状态分布：{counter_to_text(total_statuses, 6)}。',
            '',
            '## 分册导航',
            '',
        ] + table(overview_table) + [
            '',
            '## 与需求页的追踪关系',
            '',
            '- [[核心业务需求导航]]',
            '- [[后台管理需求导航]]',
            '- [[标准功能需求地图]]',
        ] + [f'- [[{title}]]' for title in overall_related[:10]] + [
            '',
            '## 观察结论',
            '',
            '- 这批历史导出资料已形成相对稳定的 12 列结构，可直接作为测试资产盘点入口。',
            '- 用例覆盖面横跨系统管理、主数据、计划管理、质量管理、制造执行、仓储物流和工装工具等多条业务线。',
            '- 历史导出中的 `编辑模式` 主要使用 `TEXT`，与导入模板样例和 skill 规则存在差异，需要在导入或生成前统一口径。',
            '- 高频标签大多可回链到现有的 MOM 需求拆解页与产品资料页，已经具备“需求 → 测试点/用例 → 模板化落地”的知识路径。',
            '',
            '## 关联页面',
            '',
            '- [[测试用例资料总览]]',
            '- [[测试用例导入模板说明]]',
            '- [[测试用例生成Skill资料包]]',
        ] + [f'- [[{title}]]' for title in workbook_titles],
    )

    template_rows = [[field, FIELD_DESCRIPTIONS.get(field, '字段说明待补充。')] for field in template_book['header']]
    template_page = GeneratedPage(
        title='测试用例导入模板说明',
        path=TEMPLATE_DIR / '测试用例导入模板说明.md',
        page_type='template',
        status='active',
        tags=['testing', 'template', 'case-design'],
        summary='说明 raw/测试资料/01_测试用例/03_测试用例导入模板 下 Excel 模板的列结构、STEP 多行录入方式和与生成 skill 的对接关系。',
        source=[template_book['file']],
        body=[
            '# 测试用例导入模板说明',
            '',
            '## 模板定位',
            '',
            f'- 来源文件：`{template_book["file"]}`。',
            f'- 工作表：`{template_book["sheet"]}`。',
            f'- 样例规模：{template_book["row_count"]} 行非表头数据，其中 4 条主用例示例通过后续空字段行演示多步骤展开。',
            f'- 列结构：共 {len(template_book["header"])} 列，不包含历史导出中的 `ID` 列。',
            '- 录入方式：样例采用 `STEP` 模式，首行填写共享字段，后续步骤行只保留 `步骤描述` 与 `预期结果`。',
            '- 追溯方式：`标签` 字段建议挂需求编号或需求主题，便于和需求页、测试点页建立链路。',
            '',
            '## 列结构',
            '',
        ] + table([['列名', '说明'], *template_rows]) + [
            '',
            '## 示例特征',
            '',
            f'- 所属模块样例：{counter_to_text(template_book["modules"], 4)}。',
            f'- 优先级样例：{counter_to_text(template_book["priorities"], 4)}。',
            f'- 责任人样例：{counter_to_text(template_book["owners"], 4)}。',
            '',
            '## 与现有知识的关系',
            '',
            '- [[测试用例模板]]：适合先在 Markdown 层整理，再映射到 Excel 导入模板。',
            '- [[需求转测试用例Skill说明]]：默认就按这个列顺序组织输出。',
            '- [[MOM产品测试用例总览]]：历史导出分册可以反向作为模板填充参考。',
            '- [[测试用例资料总览]]',
        ],
    )

    sample_page = GeneratedPage(
        title='测试用例样例说明',
        path=PROCESS_DIR / '测试用例样例说明.md',
        page_type='asset',
        status='review',
        tags=['testing', 'case-design', 'process'],
        summary='记录 raw/测试资料/01_测试用例/02_测试用例样例 目录当前为空的状态，并给出补充方向。',
        source=[rel(sample_dir)],
        body=[
            '# 测试用例样例说明',
            '',
            '## 当前状态',
            '',
            f'- 来源目录：`{rel(sample_dir)}`。',
            '- 当前未发现样例文件，属于已预留但尚未沉淀内容的目录。',
            '- 该目录适合后续补入代表性的高质量样例，用于培训、模板校验和 LLM 生成结果比对。',
            '',
            '## 建议补充内容',
            '',
            '- 一条标准主流程用例样例，展示单步和多步写法。',
            '- 一条边界/异常用例样例，展示如何拆分失败路径。',
            '- 一条跨模块联动用例样例，展示 `标签`、`备注` 与需求追溯信息。',
            '',
            '## 关联页面',
            '',
            '- [[测试用例资料总览]]',
            '- [[测试用例导入模板说明]]',
            '- [[测试用例模板]]',
            '- [[需求转测试用例Skill说明]]',
        ],
    )

    skill_bundle_page = GeneratedPage(
        title='测试用例生成Skill资料包',
        path=PROCESS_DIR / '测试用例生成Skill资料包.md',
        page_type='manual',
        status='active',
        tags=['testing', 'case-design', 'process'],
        summary='梳理 raw/测试资料/01_测试用例/生成测试用例skill 下的测试点拆分与测试用例生成资料包，并串联到导入模板。',
        source=[rel(skill_root)],
        body=[
            '# 测试用例生成Skill资料包',
            '',
            '## 资料包结构',
            '',
            '- [[需求转测试点Skill说明]]：把需求文档、原型图和流程图拆成结构化测试点。',
            '- [[需求转测试用例Skill说明]]：把测试点进一步展开为贴合本地 Excel 模板的可执行用例。',
            '- 支撑文件包括 `testcase-template.md`、`testcase-rules.md`、`testpoint-dimensions.md` 以及 MOM 补充检查清单等。',
            '',
            '## 推荐使用链路',
            '',
            '- 需求/原型/流程图 → [[需求转测试点Skill说明]] → [[需求转测试用例Skill说明]] → [[测试用例导入模板说明]] → [[MOM产品测试用例总览]]。',
            '- 如果已经有测试点，可直接跳过“测试点拆分”阶段进入用例生成。',
            '',
            '## 当前资料给出的约束',
            '',
            '- 测试点阶段强调功能流程、异常流程、边界条件、数据规则、权限控制和状态流转。',
            '- 测试用例阶段强调列结构固定、一个用例一个明确目标、多步场景拆多行。',
            '- skill 文档、模板样例和历史导出资料对 `编辑模式` 的取值存在差异，落地时要先统一平台约束。',
            '',
            '## 关联页面',
            '',
            '- [[测试用例资料总览]]',
            '- [[测试用例导入模板说明]]',
            '- [[测试用例设计]]',
            '- [[测试用例规范]]',
        ],
    )

    testcase_skill_page = GeneratedPage(
        title='需求转测试用例Skill说明',
        path=PROCESS_DIR / '需求转测试用例Skill说明.md',
        page_type='manual',
        status='active',
        tags=['testing', 'case-design', 'process'],
        summary='根据 requirement-to-testcases 资料包整理的工作流说明：从需求材料或测试点生成对齐本地 Excel 模板的结构化测试用例。',
        source=[
            rel(skill_root / 'requirement-to-testcases' / 'SKILL.md'),
            rel(skill_root / 'requirement-to-testcases' / 'assets' / 'testcase-template.md'),
            rel(skill_root / 'requirement-to-testcases' / 'references' / 'testcase-rules.md'),
        ],
        body=[
            '# 需求转测试用例Skill说明',
            '',
            '## 适用场景',
            '',
            '- 输入材料可以是需求文档、原型图、流程图，也可以是已整理好的功能测试点。',
            '- 当只有需求材料没有测试点时，先抽出功能、异常、边界、权限和状态类测试点，再继续展开用例。',
            '- 当用户要求“直接按本地测试用例模版输出”时，应优先走这套资料包。',
            '',
            '## 默认输出字段',
            '',
            '- `用例名称`、`所属模块`、`标签`、`前置条件`、`步骤描述`、`预期结果`、`编辑模式`、`备注`、`用例状态`、`责任人`、`用例等级`。',
            '- 多步用例按模板拆成多行：首行填写共享字段，后续行只保留 `步骤描述` 与 `预期结果`。',
            '- 若用户只要文本结果，可按与 Excel 同列顺序输出 Markdown 表格。',
            '',
            '## 生成规则',
            '',
            '- 一个测试用例只覆盖一个明确目标，不混塞多个验证点。',
            '- 主流程、异常流程、边界场景、权限场景和状态场景应拆开生成。',
            '- `所属模块` 默认使用 `/一级模块/二级模块/功能点` 路径写法。',
            '- `备注` 默认承载来源测试点、场景类型、验证结果和待确认信息。',
            '- 若规则不完整但仍可产出初版，应先输出，并明确标出待确认信息。',
            '',
            '## 与当前资料的关系',
            '',
            '- [[测试用例导入模板说明]]：本 skill 的目标列结构直接对齐该模板。',
            '- [[需求转测试点Skill说明]]：可作为前置步骤，先做需求分析拆分。',
            '- [[测试用例模板]]：适合在手工或半自动整理阶段先用 Markdown 形态表达。',
            '- [[MOM产品测试用例总览]]：可参考历史用例的模块路径、标签和粒度。',
            '- [[测试用例生成Skill资料包]]',
        ],
    )

    testpoint_skill_page = GeneratedPage(
        title='需求转测试点Skill说明',
        path=PROCESS_DIR / '需求转测试点Skill说明.md',
        page_type='manual',
        status='active',
        tags=['testing', 'case-design', 'process'],
        summary='根据 requirement-to-testpoints 资料包整理的需求分析说明：从需求正文、原型图或流程图拆出结构化测试点、风险和待确认项。',
        source=[
            rel(skill_root / 'requirement-to-testpoints' / 'SKILL.md'),
            rel(skill_root / 'requirement-to-testpoints' / 'references' / 'testpoint-dimensions.md'),
        ],
        body=[
            '# 需求转测试点Skill说明',
            '',
            '## 适用场景',
            '',
            '- 面向需求分析阶段，输入可以是需求文档、原型图、流程图或其组合。',
            '- 默认产出保持在“需求分析产物”层级，不直接展开为完整测试用例，除非用户明确提出。',
            '- 如果材料里能看出风险点、评审问题或跨模块影响，也应一并沉淀。',
            '',
            '## 核心拆分维度',
            '',
            '- 功能流程、异常流程、边界条件、数据规则、权限控制、状态流转、交互反馈、依赖与联动。',
            '- 面对原型图时额外关注必填项、默认值、隐藏态、禁用态和按钮可用性。',
            '- 面对流程图时同时覆盖合法流转、非法流转、回退路径和终止状态。',
            '',
            '## MOM 场景补充规则',
            '',
            '- 主数据型需求优先拆 `查询`、`新增`、`编辑`、`删除`、`引用限制` 和 `共享可见性`。',
            '- 执行型需求优先拆 `派工`、`报工`、`报检`、`授权放开` 和 `审计留痕`。',
            '- 配置型需求覆盖配置主体、匹配优先级、必填联动、最小可用配置和禁用态。',
            '- 复制或引入类需求要补充复制范围、深浅拷贝和后续同步规则。',
            '',
            '## 与当前资料的关系',
            '',
            '- [[需求转测试用例Skill说明]]：测试点确认后，下一步通常进入模板化用例生成。',
            '- [[测试用例设计]]：可以作为测试点转测试用例前的设计方法论补充。',
            '- [[功能测试检查项]]：适合作为功能维度查漏补缺的横向清单。',
            '- [[测试用例生成Skill资料包]]',
        ],
    )

    pages = [
        root_page,
        mom_overview_page,
        template_page,
        sample_page,
        skill_bundle_page,
        testcase_skill_page,
        testpoint_skill_page,
        *workbook_pages,
    ]
    for page in pages:
        write_page(page)

    replace_once(
        WIKI_ROOT / '02_测试标准&模板' / '测试标准与模板总览.md',
        '- [[可复用模板总览]]：测试计划、方案、用例、数据、报告、操作手册等模板\n- [[测试标准原始资料镜像总览]]：测试流程规范原文镜像与追溯入口',
        '- [[可复用模板总览]]：测试计划、方案、用例、数据、报告、操作手册等模板\n- [[测试用例资料总览]]：MOM 产品用例导出、导入模板、样例占位与生成 skill 的统一入口\n- [[测试标准原始资料镜像总览]]：测试流程规范原文镜像与追溯入口',
    )

    replace_once(
        WIKI_ROOT / '02_测试标准&模板' / '测试过程规范' / '测试过程规范总览.md',
        '- [[测试用例评审规程]]\n- [[回归测试清单]]\n- [[测试策略]]',
        '- [[测试用例评审规程]]\n- [[回归测试清单]]\n- [[测试策略]]\n- [[测试用例资料总览]]\n- [[MOM产品测试用例总览]]\n- [[需求转测试点Skill说明]]\n- [[需求转测试用例Skill说明]]',
    )

    replace_once(
        WIKI_ROOT / '02_测试标准&模板' / '测试过程规范' / '测试资产库总览.md',
        '- 测试用例库\n- 测试数据集\n- 回归测试清单\n- 缺陷案例库',
        '- [[测试用例资料总览]]：MOM 产品历史用例、导入模板、样例占位和生成 skill 的入口\n- 测试数据集\n- [[回归测试清单]]\n- 缺陷案例库',
    )

    replace_once(
        WIKI_ROOT / '02_测试标准&模板' / '可复用模板' / '可复用模板总览.md',
        '- [[测试用例模板]]\n- [[完整版本测试报告模板]]',
        '- [[测试用例模板]]\n- [[测试用例导入模板说明]]\n- [[完整版本测试报告模板]]',
    )

    append_section_if_missing(
        WIKI_ROOT / '02_测试标准&模板' / '可复用模板' / '测试用例模板.md',
        '## 关联页面',
        ['- [[测试用例导入模板说明]]', '- [[需求转测试用例Skill说明]]', '- [[测试用例资料总览]]'],
    )


if __name__ == '__main__':
    main()
